"""
harvester.py — Stats SA CPI Web Harvester.

Ingests the official South African headline Consumer Price Index ("Total
country", all items) time series directly from Statistics South Africa's
own published time-series archive, and writes the requested trailing window
of months to data/stats_sa_cpi_archive.csv in the Date (YYYY-MM), CPI_Value
schema that core/validator.py expects.

Source
------
Stats SA publishes the full CPI(COICOP) time series (Jan 2008 → latest
release month) as a single cumulative Excel workbook, refreshed every month
alongside the P0141 CPI statistical release:

    Listing page : https://www.statssa.gov.za/?page_id=1847
    File pattern  : https://www.statssa.gov.za/timeseriesdata/Excel/
                    P0141 - CPI(COICOP) from Jan 2008 (YYYYMM).zip

Inside the workbook, series code CPI60001 / classification "Total country"
is the headline all-items, whole-country index — the same series Stats SA
reports as "the CPI" in the P0141 release. Its base period is stated in the
row itself (currently "Dec 2024 = 100", following Stats SA's 2025 rebasing).

Usage
-----
    python core/harvester.py                     # ingest latest 24 months
    python core/harvester.py --months-back 12     # ingest latest 12 months
    python core/harvester.py --out data/stats_sa_cpi_archive.csv

Design notes
------------
- The listing page is scraped for the current release tag (YYYYMM) instead
  of hardcoding a month, so this keeps working as Stats SA publishes new
  releases without code changes.
- The archive this writes is a POINT-IN-TIME CSV: each run replaces the
  trailing window, it never mutates a historical month's value in place —
  consistent with core/validator.py Stage 1a (Bitemporal Core / PIT).
- This intentionally does NOT run validator.py itself. Run
  `python core/validator.py` (or the Streamlit app) afterwards to gate the
  freshly ingested file through the 10-stage pipeline before trusting it.
"""
from __future__ import annotations

import argparse
import csv
import io
import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

LISTING_PAGE = "https://www.statssa.gov.za/?page_id=1847"
ZIP_URL_TEMPLATE = (
    "https://www.statssa.gov.za/timeseriesdata/Excel/"
    "P0141 - CPI(COICOP) from Jan 2008 ({tag}).zip"
)
RELEASE_TAG_PATTERN = re.compile(r"CPI\(COICOP\) from Jan 2008 \((\d{6})\)\.zip", re.IGNORECASE)

HEADLINE_SERIES_CODE = "CPI60001"
HEADLINE_SERIES_LABEL = "Total country"
USER_AGENT = "Mozilla/5.0 (LekwankwaSCMEngine/1.0; +https://github.com/)"


def discover_latest_release_tag(timeout: int = 30) -> str:
    """Scrape the Stats SA time-series listing page for the current release tag (YYYYMM)."""
    resp = requests.get(LISTING_PAGE, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    resp.raise_for_status()
    match = RELEASE_TAG_PATTERN.search(resp.text)
    if not match:
        raise RuntimeError(
            "Could not find a 'CPI(COICOP) from Jan 2008 (YYYYMM).zip' link on the "
            f"Stats SA time-series listing page ({LISTING_PAGE}). The page layout may "
            "have changed — pass --release-tag explicitly to bypass discovery."
        )
    return match.group(1)


def download_workbook(release_tag: str, timeout: int = 60) -> openpyxl.Workbook:
    """Download the CPI(COICOP) zip for a given release tag and return the parsed workbook."""
    url = ZIP_URL_TEMPLATE.format(tag=release_tag)
    resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    resp.raise_for_status()

    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise RuntimeError(f"No .xlsx file found inside {url}")
        with zf.open(xlsx_names[0]) as f:
            return openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)


def extract_headline_series(wb: openpyxl.Workbook) -> dict:
    """Return {'YYYY-MM': cpi_value} for the headline 'Total country' all-items series."""
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    month_col_positions = [
        (i, h) for i, h in enumerate(header) if isinstance(h, str) and re.fullmatch(r"MO\d{6}", h)
    ]
    if not month_col_positions:
        raise RuntimeError("No monthly 'MOxxYYYY' columns found in the CPI workbook header row.")

    for row in rows:
        code, label = row[2], row[6]
        if code == HEADLINE_SERIES_CODE and label == HEADLINE_SERIES_LABEL:
            series = {}
            for idx, col_name in month_col_positions:
                m, y = int(col_name[2:4]), int(col_name[4:8])
                value = row[idx]
                if value is not None:
                    series[f"{y:04d}-{m:02d}"] = float(value)
            return series

    raise RuntimeError(
        f"Headline series {HEADLINE_SERIES_CODE} ('{HEADLINE_SERIES_LABEL}') not found "
        "in the CPI workbook — Stats SA may have changed the series code."
    )


def trailing_window(series: dict, months_back: int) -> list[tuple[str, float]]:
    """Sort the series ascending by month and keep only the last `months_back` entries."""
    ordered = sorted(series.items())
    return ordered[-months_back:] if months_back else ordered


def write_archive_csv(rows: list[tuple[str, float]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "CPI_Value"])
        writer.writerows(rows)


def ingest(months_back: int = 24, out_path: Path = None, release_tag: str = None) -> Path:
    out_path = out_path or Path("data") / "stats_sa_cpi_archive.csv"
    release_tag = release_tag or discover_latest_release_tag()

    print(f"[harvester] Release tag discovered : {release_tag}")
    wb = download_workbook(release_tag)

    print(f"[harvester] Extracting headline series {HEADLINE_SERIES_CODE} ('{HEADLINE_SERIES_LABEL}')...")
    series = extract_headline_series(wb)
    print(f"[harvester] Full series length      : {len(series)} months "
          f"({min(series)} .. {max(series)})")

    rows = trailing_window(series, months_back)
    print(f"[harvester] Writing trailing window  : {rows[0][0]} .. {rows[-1][0]} "
          f"({len(rows)} months) -> {out_path}")
    write_archive_csv(rows, out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--months-back", type=int, default=24,
                         help="How many trailing months to write into the archive (default: 24).")
    parser.add_argument("--out", type=Path, default=Path("data") / "stats_sa_cpi_archive.csv",
                         help="Output CSV path (default: data/stats_sa_cpi_archive.csv).")
    parser.add_argument("--release-tag", type=str, default=None,
                         help="Override auto-discovery with an explicit YYYYMM release tag "
                              "(e.g. 202606). Useful if the listing page layout changes.")
    args = parser.parse_args()

    ingest(months_back=args.months_back, out_path=args.out, release_tag=args.release_tag)
    print(f"[harvester] Done at {datetime.now().isoformat()}. "
          f"Run `python core/validator.py` next to gate the file through the 10-stage pipeline.")


if __name__ == "__main__":
    main()
